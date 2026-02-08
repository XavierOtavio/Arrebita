from decimal import Decimal
from io import BytesIO
import datetime as dt
import uuid

from django.db import connection
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook, load_workbook

from Accounts.models import User
from Events.models import EventListView

EVENT_STATUS_LABELS = {
    "draft": "Rascunho",
    "published": "Publicado",
    "cancelled": "Cancelado",
    "archived": "Arquivado",
}


def dashboard(request):
    return render(request, "dashboard.html")


def dictfetchall(cursor):
    cols = [col[0] for col in cursor.description]
    return [dict(zip(cols, row)) for row in cursor.fetchall()]


def _format_event_price(event):
    if event.is_free:
        return "Gratuito"
    if event.price_cents is None:
        return "Preco a definir"
    currency = (event.currency_code or "EUR").strip() or "EUR"
    return f"{currency} {event.price_cents / 100:.2f}"


def _format_event_location(event):
    if event.is_online:
        return "Online"
    parts = []
    if event.venue_name:
        parts.append(event.venue_name)
    city_parts = [event.city, event.region, event.country_code]
    city_line = ", ".join([part for part in city_parts if part])
    if city_line:
        parts.append(city_line)
    return " - ".join(parts)


def _events_queryset_from_request(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    timing = (request.GET.get("timing") or "").strip()
    mode = (request.GET.get("mode") or "").strip()
    sort = (request.GET.get("sort") or "").strip()

    events_qs = EventListView.objects.all()

    if q:
        events_qs = events_qs.filter(
            Q(title__icontains=q)
            | Q(slug__icontains=q)
            | Q(city__icontains=q)
            | Q(venue_name__icontains=q)
        )

    if status:
        events_qs = events_qs.filter(status=status)

    if timing == "upcoming":
        events_qs = events_qs.filter(is_upcoming=True)
    elif timing == "finished":
        events_qs = events_qs.filter(is_finished=True)
    elif timing == "ongoing":
        events_qs = events_qs.filter(is_upcoming=False, is_finished=False)

    if mode == "online":
        events_qs = events_qs.filter(is_online=True)
    elif mode == "onsite":
        events_qs = events_qs.filter(is_online=False)

    allowed_sorts = {
        "starts_at": "starts_at",
        "-starts_at": "-starts_at",
        "price": "price_cents",
        "-price": "-price_cents",
        "created_at": "created_at",
        "-created_at": "-created_at",
    }
    events_qs = events_qs.order_by(allowed_sorts.get(sort, "-starts_at"))

    filters = {
        "q": q,
        "status": status,
        "timing": timing,
        "mode": mode,
        "sort": sort,
    }
    return events_qs, filters


def backoffice_wines(request):
    """
    Lista de vinhos do backoffice, baseada na view vw_wine_list.
    A view deve expor, pelo menos:
      - wine_id, sku, name, type_id, type_label
      - region (nome da região)
      - vintage_year, price, stock_qty
      - tasting_notes, alcohol_content, serving_temperature,
        bottle_capacity, pairing, winemaker
      - promo_pct_off, promo_price, has_active_promo
    """

    # 1) Tipos de vinho
    with connection.cursor() as cur:
        cur.execute("SELECT * FROM get_wine_types();")
        wine_types = dictfetchall(cur)

    # 2) Regiões normalizadas
    with connection.cursor() as cur:
        cur.execute(
            "SELECT region_id, name FROM public.regions ORDER BY name;"
        )
        regions = dictfetchall(cur)

    # 3) Filtros da querystring
    q = request.GET.get("q") or ""
    type_id = request.GET.get("type") or None
    region = request.GET.get("region") or ""
    only_on_promo = bool(request.GET.get("only_on_promo"))

    # 4) Query base
    sql = """
          SELECT
              wine_id,
              sku,
              name,
              type_id,
              type_label,
              region,
              vintage_year,
              price,
              stock_qty,
              tasting_notes,
              alcohol_content,
              serving_temperature,
              bottle_capacity,
              pairing,
              winemaker,
              promo_pct_off,
              promo_price,
              has_active_promo
          FROM public.vw_wine_list
          WHERE 1 = 1
          """
    params = []

    # 5) Filtro de pesquisa geral (nome / SKU)
    if q:
        sql += " AND (name ILIKE %s OR sku ILIKE %s)"
        params.extend([f"%{q}%", f"%{q}%"])

    # 6) Filtro por tipo de vinho
    if type_id:
        sql += " AND type_id = %s"
        params.append(uuid.UUID(type_id))

    # 7) Filtro por nome da região (exposto pela view)
    if region:
        sql += " AND region ILIKE %s"
        params.append(f"%{region}%")

    # 8) Filtro: apenas vinhos com promoção ativa
    if only_on_promo:
        sql += " AND has_active_promo = TRUE"

    # 9) Execução da query de vinhos
    with connection.cursor() as cur:
        cur.execute(sql, params)
        wines = dictfetchall(cur)

    # 10) Carregar imagens de vinhos e associar a cada vinho
    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT image_id, wine_id, image_url, image_type, created_at
            FROM public.wine_images
            ORDER BY created_at DESC;
            """
        )
        images = dictfetchall(cur)

    images_by_wine = {}
    for img in images:
        wkey = str(img["wine_id"])
        images_by_wine.setdefault(wkey, []).append(img)

    for w in wines:
        w["images"] = images_by_wine.get(str(w["wine_id"]), [])

    context = {
        "wines": wines,
        "wine_types": wine_types,
        "regions": regions,
    }
    return render(request, "vinhos_catalogo.html", context)


def backoffice_wine_create(request):
    if request.method != "POST":
        return redirect(reverse("backoffice:backoffice_wines"))

    data = request.POST

    def to_int(value):
        value = value or None
        return int(value) if value is not None and value != "" else None

    def to_float(value):
        value = value or None
        return float(value) if value is not None and value != "" else None

    region_id_str = data.get("region_id")
    region_id = uuid.UUID(region_id_str) if region_id_str else None

    params = [
        data.get("sku"),
        data.get("name"),
        uuid.UUID(data.get("type_id")),
        region_id,                              # region_id (UUID)
        to_int(data.get("vintage_year")),
        to_float(data.get("price")),
        to_int(data.get("stock_qty")),
        data.get("tasting_notes") or None,
        to_float(data.get("alcohol_content")),
        to_float(data.get("serving_temperature")),
        to_float(data.get("bottle_capacity")),
        data.get("pairing") or None,
        data.get("winemaker") or None,
        ]

    with connection.cursor() as cur:
        cur.execute(
            """
            CALL public.create_wine(
                %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s
            );
            """,
            params,
        )

    return redirect(reverse("backoffice:backoffice_wines"))


def backoffice_wine_update(request, wine_id):
    """
    Atualiza um vinho existente, chamado a partir do modal Editar.
    """
    if request.method != "POST":
        return redirect(reverse("backoffice:backoffice_wines"))

    data = request.POST

    def to_int(value):
        value = value or None
        return int(value) if value is not None and value != "" else None

    def to_float(value):
        value = value or None
        return float(value) if value is not None and value != "" else None

    region_id_str = data.get("region_id")
    region_id = uuid.UUID(region_id_str) if region_id_str else None

    params = [
        uuid.UUID(wine_id),
        data.get("sku"),
        data.get("name"),
        uuid.UUID(data.get("type_id")),
        region_id,                              # region_id (UUID)
        to_int(data.get("vintage_year")),
        to_float(data.get("price")),
        to_int(data.get("stock_qty")),
        data.get("tasting_notes") or None,
        to_float(data.get("alcohol_content")),
        to_float(data.get("serving_temperature")),
        to_float(data.get("bottle_capacity")),
        data.get("pairing") or None,
        data.get("winemaker") or None,
        ]

    with connection.cursor() as cur:
        cur.execute(
            """
            CALL public.update_wine(
                %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s
            );
            """,
            params,
        )

    return redirect(reverse("backoffice:backoffice_wines"))


def backoffice_wine_delete(request, wine_id):
    if request.method != "POST":
        return redirect(reverse("backoffice:backoffice_wines"))

    with connection.cursor() as cur:
        cur.execute(
            "CALL public.delete_wine(%s);",
            [uuid.UUID(wine_id)],
        )

    return redirect(reverse("backoffice:backoffice_wines"))


def backoffice_wine_image_create(request, wine_id):
    """
    Cria uma imagem para um vinho específico.
    """
    if request.method != "POST":
        return redirect(reverse("backoffice:backoffice_wines"))

    data = request.POST
    image_url = data.get("image_url") or ""
    image_type = data.get("image_type") or "catalog"

    if not image_url.strip():
        # Não vale a pena criar entradas vazias
        return redirect(reverse("backoffice:backoffice_wines"))

    with connection.cursor() as cur:
        cur.execute(
            """
            CALL public.create_wine_image(%s, %s, %s);
            """,
            [uuid.UUID(wine_id), image_url, image_type],
        )

    return redirect(reverse("backoffice:backoffice_wines"))


def backoffice_wine_image_delete(request, wine_id, image_id):
    """
    Apaga uma imagem de vinho.
    """
    if request.method != "POST":
        return redirect(reverse("backoffice:backoffice_wines"))

    with connection.cursor() as cur:
        cur.execute(
            "CALL public.delete_wine_image(%s);",
            [uuid.UUID(image_id)],
        )

    return redirect(reverse("backoffice:backoffice_wines"))


def backoffice_events(request):
    events_qs, filters = _events_queryset_from_request(request)

    events = list(events_qs[:200])
    for event in events:
        title = (event.title or "").strip()
        slug = (event.slug or "").strip()
        if title:
            event.display_title = title
        elif slug:
            event.display_title = slug.replace("-", " ").title()
        else:
            event.display_title = f"Evento {event.event_id}"

        event.format_label = "Online" if event.is_online else "Presencial"
        event.price_display = _format_event_price(event)
        event.location_display = _format_event_location(event)
        event.status_label = EVENT_STATUS_LABELS.get(event.status, event.status or "")
        if event.price_cents is not None:
            event.price_eur = f"{event.price_cents / 100:.2f}"
        else:
            event.price_eur = ""

        if event.is_finished:
            event.timing_label = "Terminado"
        elif event.is_upcoming:
            event.timing_label = "Proximo"
        else:
            event.timing_label = "Em curso"

    params = request.GET.copy()
    export_querystring = params.urlencode()

    context = {
        "events": events,
        "export_querystring": export_querystring,
        **filters,
    }
    return render(request, "events_catalogo.html", context)


def backoffice_event_create(request):
    if request.method != "POST":
        return redirect(reverse("backoffice:backoffice_events"))

    data = request.POST

    def to_int(value):
        value = (value or "").strip()
        return int(value) if value else None

    def to_float(value):
        value = (value or "").strip()
        return float(value) if value else None

    title = (data.get("title") or "").strip()
    slug = (data.get("slug") or "").strip()
    if not slug and title:
        slug = slugify(title)

    if not title or not slug:
        return redirect(reverse("backoffice:backoffice_events"))

    summary = (data.get("summary") or "").strip() or None
    description = (data.get("description") or "").strip() or None

    is_online = data.get("is_online") == "on"
    online_url = (data.get("online_url") or "").strip() or None
    if not is_online:
        online_url = None

    venue_name = (data.get("venue_name") or "").strip() or None
    address_line1 = (data.get("address_line1") or "").strip() or None
    address_line2 = (data.get("address_line2") or "").strip() or None
    postal_code = (data.get("postal_code") or "").strip() or None
    city = (data.get("city") or "").strip() or None
    region = (data.get("region") or "").strip() or None
    country_code = (data.get("country_code") or "").strip() or None

    latitude = to_float(data.get("latitude"))
    longitude = to_float(data.get("longitude"))

    starts_at = (data.get("starts_at") or "").strip()
    ends_at = (data.get("ends_at") or "").strip() or None
    timezone_name = (data.get("timezone") or "").strip() or "Europe/Lisbon"

    capacity = to_int(data.get("capacity"))
    is_free = data.get("is_free") == "on"

    price_value = to_float(data.get("price_eur"))
    price_cents = int(round(price_value * 100)) if price_value is not None else None
    currency_code = (data.get("currency_code") or "EUR").strip() or "EUR"

    if is_free:
        price_cents = None
        currency_code = None

    status = (data.get("status") or "draft").strip()
    published_at = timezone.now() if status == "published" else None

    if not starts_at:
        return redirect(reverse("backoffice:backoffice_events"))

    if is_online and not online_url:
        return redirect(reverse("backoffice:backoffice_events"))

    with connection.cursor() as cur:
        cur.execute(
            """
            INSERT INTO public.events (
                title,
                slug,
                summary,
                description,
                is_online,
                online_url,
                venue_name,
                address_line1,
                address_line2,
                postal_code,
                city,
                region,
                country_code,
                latitude,
                longitude,
                starts_at,
                ends_at,
                timezone,
                capacity,
                price_cents,
                currency_code,
                is_free,
                status,
                published_at
            ) VALUES (
                %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s
            );
            """,
            [
                title,
                slug,
                summary,
                description,
                is_online,
                online_url,
                venue_name,
                address_line1,
                address_line2,
                postal_code,
                city,
                region,
                country_code,
                latitude,
                longitude,
                starts_at,
                ends_at,
                timezone_name,
                capacity,
                price_cents,
                currency_code,
                is_free,
                status,
                published_at,
            ],
        )
        cur.execute("REFRESH MATERIALIZED VIEW public.mv_events_all;")

    return redirect(reverse("backoffice:backoffice_events"))


def backoffice_event_update(request, event_id):
    if request.method != "POST":
        return redirect(reverse("backoffice:backoffice_events"))

    data = request.POST

    def to_int(value):
        value = (value or "").strip()
        return int(value) if value else None

    def to_float(value):
        value = (value or "").strip()
        return float(value) if value else None

    title = (data.get("title") or "").strip()
    slug = (data.get("slug") or "").strip()
    if not slug and title:
        slug = slugify(title)

    if not title or not slug:
        return redirect(reverse("backoffice:backoffice_events"))

    summary = (data.get("summary") or "").strip() or None
    description = (data.get("description") or "").strip() or None

    is_online = data.get("is_online") == "on"
    online_url = (data.get("online_url") or "").strip() or None
    if not is_online:
        online_url = None

    venue_name = (data.get("venue_name") or "").strip() or None
    address_line1 = (data.get("address_line1") or "").strip() or None
    address_line2 = (data.get("address_line2") or "").strip() or None
    postal_code = (data.get("postal_code") or "").strip() or None
    city = (data.get("city") or "").strip() or None
    region = (data.get("region") or "").strip() or None
    country_code = (data.get("country_code") or "").strip().upper() or None

    latitude = to_float(data.get("latitude"))
    longitude = to_float(data.get("longitude"))

    starts_at = (data.get("starts_at") or "").strip()
    ends_at = (data.get("ends_at") or "").strip() or None
    timezone_name = (data.get("timezone") or "").strip() or "Europe/Lisbon"

    capacity = to_int(data.get("capacity"))
    is_free = data.get("is_free") == "on"

    price_value = to_float(data.get("price_eur"))
    price_cents = int(round(price_value * 100)) if price_value is not None else None
    currency_code = (data.get("currency_code") or "EUR").strip() or "EUR"

    if is_free:
        price_cents = None
        currency_code = None

    status = (data.get("status") or "draft").strip()

    if not starts_at:
        return redirect(reverse("backoffice:backoffice_events"))

    if is_online and not online_url:
        return redirect(reverse("backoffice:backoffice_events"))

    with connection.cursor() as cur:
        cur.execute(
            """
            UPDATE public.events
            SET title = %s,
                slug = %s,
                summary = %s,
                description = %s,
                is_online = %s,
                online_url = %s,
                venue_name = %s,
                address_line1 = %s,
                address_line2 = %s,
                postal_code = %s,
                city = %s,
                region = %s,
                country_code = %s,
                latitude = %s,
                longitude = %s,
                starts_at = %s,
                ends_at = %s,
                timezone = %s,
                capacity = %s,
                price_cents = %s,
                currency_code = %s,
                is_free = %s,
                status = %s,
                published_at = CASE
                    WHEN %s = 'published' THEN COALESCE(published_at, now())
                    ELSE NULL
                END,
                updated_at = now()
            WHERE event_id = %s;
            """,
            [
                title,
                slug,
                summary,
                description,
                is_online,
                online_url,
                venue_name,
                address_line1,
                address_line2,
                postal_code,
                city,
                region,
                country_code,
                latitude,
                longitude,
                starts_at,
                ends_at,
                timezone_name,
                capacity,
                price_cents,
                currency_code,
                is_free,
                status,
                status,
                event_id,
            ],
        )
        cur.execute("REFRESH MATERIALIZED VIEW public.mv_events_all;")

    return redirect(reverse("backoffice:backoffice_events"))


def backoffice_events_export(request):
    events_qs, _filters = _events_queryset_from_request(request)
    events = list(events_qs[:2000])

    headers = [
        "event_id",
        "title",
        "slug",
        "summary",
        "description",
        "is_online",
        "online_url",
        "venue_name",
        "address_line1",
        "address_line2",
        "postal_code",
        "city",
        "region",
        "country_code",
        "latitude",
        "longitude",
        "starts_at",
        "ends_at",
        "timezone",
        "capacity",
        "price_cents",
        "price_eur",
        "currency_code",
        "is_free",
        "status",
        "published_at",
        "created_at",
        "updated_at",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"
    ws.append(headers)

    def to_excel_datetime(value):
        if not value:
            return None
        if timezone.is_aware(value):
            value = timezone.localtime(value)
            return value.replace(tzinfo=None)
        return value

    for event in events:
        price_eur = f"{event.price_cents / 100:.2f}" if event.price_cents is not None else ""
        row = [
            str(event.event_id),
            event.title,
            event.slug,
            event.summary,
            event.description,
            event.is_online,
            event.online_url,
            event.venue_name,
            event.address_line1,
            event.address_line2,
            event.postal_code,
            event.city,
            event.region,
            event.country_code,
            float(event.latitude) if isinstance(event.latitude, Decimal) else event.latitude,
            float(event.longitude) if isinstance(event.longitude, Decimal) else event.longitude,
            to_excel_datetime(event.starts_at),
            to_excel_datetime(event.ends_at),
            event.timezone,
            event.capacity,
            event.price_cents,
            price_eur,
            event.currency_code,
            event.is_free,
            event.status,
            to_excel_datetime(event.published_at),
            to_excel_datetime(event.created_at),
            to_excel_datetime(event.updated_at),
        ]
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="events-export.xlsx"'
    return response


def backoffice_events_import(request):
    if request.method != "POST":
        return redirect(reverse("backoffice:backoffice_events"))

    upload = request.FILES.get("events_file")
    if not upload:
        return redirect(reverse("backoffice:backoffice_events"))

    wb = load_workbook(upload, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return redirect(reverse("backoffice:backoffice_events"))

    def normalize_header(value):
        if value is None:
            return ""
        return str(value).strip().lower().replace(" ", "_")

    headers = [normalize_header(h) for h in rows[0]]
    header_map = {name: idx for idx, name in enumerate(headers) if name}

    def cell(row, key):
        idx = header_map.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    def parse_bool(value):
        if isinstance(value, bool):
            return value
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return value != 0
        value = str(value).strip().lower()
        return value in {"1", "true", "yes", "y", "sim"}

    def parse_datetime(value):
        if value is None or value == "":
            return None
        if isinstance(value, dt.datetime):
            return value
        if isinstance(value, dt.date):
            return dt.datetime.combine(value, dt.time(0, 0))
        text = str(value).strip()
        try:
            return dt.datetime.fromisoformat(text)
        except ValueError:
            return None

    def to_float(value):
        if value in ("", None):
            return None
        if isinstance(value, (int, float, Decimal)):
            return float(value)
        text = str(value).strip().replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None

    def parse_uuid(value):
        if value in ("", None):
            return None
        try:
            return uuid.UUID(str(value))
        except (ValueError, TypeError):
            return None

    for row in rows[1:]:
        if row is None or all(cell is None for cell in row):
            continue

        title = (cell(row, "title") or "").strip()
        slug = (cell(row, "slug") or "").strip()
        if not slug and title:
            slug = slugify(title)

        if not title or not slug:
            continue

        summary = (cell(row, "summary") or "").strip() or None
        description = (cell(row, "description") or "").strip() or None

        event_id = parse_uuid(cell(row, "event_id"))

        is_online = parse_bool(cell(row, "is_online"))
        online_url = (cell(row, "online_url") or "").strip() or None
        if not is_online:
            online_url = None

        if is_online and not online_url:
            continue

        venue_name = (cell(row, "venue_name") or "").strip() or None
        address_line1 = (cell(row, "address_line1") or "").strip() or None
        address_line2 = (cell(row, "address_line2") or "").strip() or None
        postal_code = (cell(row, "postal_code") or "").strip() or None
        city = (cell(row, "city") or "").strip() or None
        region = (cell(row, "region") or "").strip() or None
        country_code = (cell(row, "country_code") or "").strip().upper() or None

        latitude = to_float(cell(row, "latitude"))
        longitude = to_float(cell(row, "longitude"))

        starts_at = parse_datetime(cell(row, "starts_at"))
        ends_at = parse_datetime(cell(row, "ends_at"))
        timezone_name = (cell(row, "timezone") or "").strip() or "Europe/Lisbon"

        capacity_value = to_float(cell(row, "capacity"))
        capacity = int(capacity_value) if capacity_value is not None else None

        is_free = parse_bool(cell(row, "is_free"))

        price_cents = to_float(cell(row, "price_cents"))
        if price_cents is None:
            price_eur = to_float(cell(row, "price_eur"))
            if price_eur is not None:
                price_cents = int(round(price_eur * 100))
            else:
                price_cents = None
        else:
            price_cents = int(round(price_cents))

        currency_code = (cell(row, "currency_code") or "EUR").strip() or "EUR"
        if is_free:
            price_cents = None
            currency_code = None

        status = (cell(row, "status") or "draft").strip()

        if not starts_at:
            continue

        published_at = timezone.now() if status == "published" else None

        params_common = [
            title,
            slug,
            summary,
            description,
            is_online,
            online_url,
            venue_name,
            address_line1,
            address_line2,
            postal_code,
            city,
            region,
            country_code,
            latitude,
            longitude,
            starts_at,
            ends_at,
            timezone_name,
            capacity,
            price_cents,
            currency_code,
            is_free,
            status,
            published_at,
        ]

        with connection.cursor() as cur:
            if event_id:
                cur.execute(
                    """
                    UPDATE public.events
                    SET title = %s,
                        slug = %s,
                        summary = %s,
                        description = %s,
                        is_online = %s,
                        online_url = %s,
                        venue_name = %s,
                        address_line1 = %s,
                        address_line2 = %s,
                        postal_code = %s,
                        city = %s,
                        region = %s,
                        country_code = %s,
                        latitude = %s,
                        longitude = %s,
                        starts_at = %s,
                        ends_at = %s,
                        timezone = %s,
                        capacity = %s,
                        price_cents = %s,
                        currency_code = %s,
                        is_free = %s,
                        status = %s,
                        published_at = CASE
                            WHEN %s = 'published' THEN COALESCE(published_at, now())
                            ELSE NULL
                        END,
                        updated_at = now()
                    WHERE event_id = %s;
                    """,
                    [
                        *params_common[:-1],
                        status,
                        event_id,
                    ],
                )
                if cur.rowcount == 0:
                    cur.execute(
                        """
                        INSERT INTO public.events (
                            event_id,
                            title,
                            slug,
                            summary,
                            description,
                            is_online,
                            online_url,
                            venue_name,
                            address_line1,
                            address_line2,
                            postal_code,
                            city,
                            region,
                            country_code,
                            latitude,
                            longitude,
                            starts_at,
                            ends_at,
                            timezone,
                            capacity,
                            price_cents,
                            currency_code,
                            is_free,
                            status,
                            published_at
                        ) VALUES (
                            %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s
                        );
                        """,
                        [event_id, *params_common],
                    )
            else:
                cur.execute(
                    """
                    UPDATE public.events
                    SET title = %s,
                        summary = %s,
                        description = %s,
                        is_online = %s,
                        online_url = %s,
                        venue_name = %s,
                        address_line1 = %s,
                        address_line2 = %s,
                        postal_code = %s,
                        city = %s,
                        region = %s,
                        country_code = %s,
                        latitude = %s,
                        longitude = %s,
                        starts_at = %s,
                        ends_at = %s,
                        timezone = %s,
                        capacity = %s,
                        price_cents = %s,
                        currency_code = %s,
                        is_free = %s,
                        status = %s,
                        published_at = CASE
                            WHEN %s = 'published' THEN COALESCE(published_at, now())
                            ELSE NULL
                        END,
                        updated_at = now()
                    WHERE slug = %s;
                    """,
                    [
                        title,
                        summary,
                        description,
                        is_online,
                        online_url,
                        venue_name,
                        address_line1,
                        address_line2,
                        postal_code,
                        city,
                        region,
                        country_code,
                        latitude,
                        longitude,
                        starts_at,
                        ends_at,
                        timezone_name,
                        capacity,
                        price_cents,
                        currency_code,
                        is_free,
                        status,
                        status,
                        slug,
                    ],
                )
                if cur.rowcount == 0:
                    cur.execute(
                        """
                        INSERT INTO public.events (
                            title,
                            slug,
                            summary,
                            description,
                            is_online,
                            online_url,
                            venue_name,
                            address_line1,
                            address_line2,
                            postal_code,
                            city,
                            region,
                            country_code,
                            latitude,
                            longitude,
                            starts_at,
                            ends_at,
                            timezone,
                            capacity,
                            price_cents,
                            currency_code,
                            is_free,
                            status,
                            published_at
                        ) VALUES (
                            %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s
                        );
                        """,
                        params_common,
                    )

    with connection.cursor() as cur:
        cur.execute("REFRESH MATERIALIZED VIEW public.mv_events_all;")

    return redirect(reverse("backoffice:backoffice_events"))


def backoffice_event_delete(request, event_id):
    if request.method != "POST":
        return redirect(reverse("backoffice:backoffice_events"))

    with connection.cursor() as cur:
        cur.execute(
            "DELETE FROM public.events WHERE event_id = %s;",
            [event_id],
        )
        cur.execute("REFRESH MATERIALIZED VIEW public.mv_events_all;")

    return redirect(reverse("backoffice:backoffice_events"))


def backoffice_users(request):
    q = (request.GET.get("q") or "").strip()
    role = (request.GET.get("role") or "").strip()

    users_qs = User.objects.all()
    if q:
        users_qs = users_qs.filter(Q(full_name__icontains=q) | Q(email__icontains=q))
    if role:
        users_qs = users_qs.filter(role=role)

    users = list(users_qs.order_by("user_id")[:500])
    roles = list(User.objects.values_list("role", flat=True).distinct().order_by("role"))

    context = {
        "users": users,
        "roles": roles,
        "q": q,
        "role": role,
    }
    return render(request, "users_catalogo.html", context)


def _now_naive():
    now = timezone.now()
    if timezone.is_aware(now):
        now = timezone.localtime(now).replace(tzinfo=None)
    return now


def backoffice_user_create(request):
    if request.method != "POST":
        return redirect(reverse("backoffice:backoffice_users"))

    data = request.POST
    email = (data.get("email") or "").strip()
    full_name = (data.get("full_name") or "").strip()
    role = (data.get("role") or "").strip()
    password_hash = (data.get("password_hash") or "").strip()

    if not email or not full_name or not role or not password_hash:
        return redirect(reverse("backoffice:backoffice_users"))

    with connection.cursor() as cur:
        cur.execute(
            """
            INSERT INTO public.users (email, password_hash, full_name, role, created_at)
            VALUES (%s, %s, %s, %s, %s);
            """,
            [email, password_hash, full_name, role, _now_naive()],
        )

    return redirect(reverse("backoffice:backoffice_users"))


def backoffice_user_update(request, user_id):
    if request.method != "POST":
        return redirect(reverse("backoffice:backoffice_users"))

    data = request.POST
    email = (data.get("email") or "").strip()
    full_name = (data.get("full_name") or "").strip()
    role = (data.get("role") or "").strip()
    password_hash = (data.get("password_hash") or "").strip()

    if not email or not full_name or not role:
        return redirect(reverse("backoffice:backoffice_users"))

    if password_hash:
        sql = """
            UPDATE public.users
            SET email = %s,
                password_hash = %s,
                full_name = %s,
                role = %s
            WHERE user_id = %s;
        """
        params = [email, password_hash, full_name, role, user_id]
    else:
        sql = """
            UPDATE public.users
            SET email = %s,
                full_name = %s,
                role = %s
            WHERE user_id = %s;
        """
        params = [email, full_name, role, user_id]

    with connection.cursor() as cur:
        cur.execute(sql, params)

    return redirect(reverse("backoffice:backoffice_users"))


def backoffice_user_access(request):
    users = list(User.objects.order_by("full_name"))
    selected_user_id = request.GET.get("user_id") or request.POST.get("user_id")
    selected_user = None

    if selected_user_id:
        try:
            selected_user = User.objects.get(user_id=int(selected_user_id))
        except (User.DoesNotExist, ValueError, TypeError):
            selected_user = None

    permission_items = []
    role_permissions = set()

    with connection.cursor() as cur:
        cur.execute(
            """
            SELECT permission, MAX(description) AS description
            FROM public.role_permissions
            GROUP BY permission
            ORDER BY permission;
            """
        )
        permission_items = [
            {"permission": perm, "description": desc}
            for perm, desc in cur.fetchall()
        ]

        if selected_user:
            cur.execute(
                "SELECT permission FROM public.role_permissions WHERE role = %s ORDER BY permission;",
                [selected_user.role],
            )
            role_permissions = {row[0] for row in cur.fetchall()}

    if request.method == "POST" and selected_user:
        delete_permission = (request.POST.get("delete_permission") or "").strip()
        if delete_permission:
            with connection.cursor() as cur:
                cur.execute(
                    "DELETE FROM public.role_permissions WHERE permission = %s;",
                    [delete_permission],
                )
            return redirect(f"{request.path}?user_id={selected_user.user_id}")

        selected = set(request.POST.getlist("permissions"))
        new_permission = (request.POST.get("new_permission") or "").strip()
        new_description = (request.POST.get("new_description") or "").strip() or None

        perm_list = request.POST.getlist("perm_list")
        perm_desc = request.POST.getlist("perm_desc")
        desc_map = {}
        for perm, desc in zip(perm_list, perm_desc):
            desc_map[perm] = (desc or "").strip() or None

        if new_permission:
            selected.add(new_permission)
            if new_description:
                desc_map[new_permission] = new_description

        with connection.cursor() as cur:
            for perm, desc in desc_map.items():
                if desc:
                    cur.execute(
                        "UPDATE public.role_permissions SET description = %s WHERE permission = %s;",
                        [desc, perm],
                    )

            cur.execute("DELETE FROM public.role_permissions WHERE role = %s;", [selected_user.role])
            for perm in sorted(selected):
                desc = desc_map.get(perm)
                if not desc:
                    for item in permission_items:
                        if item["permission"] == perm:
                            desc = item["description"]
                            break
                cur.execute(
                    """
                    INSERT INTO public.role_permissions (role, permission, description)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (role, permission) DO UPDATE
                    SET description = COALESCE(EXCLUDED.description, role_permissions.description);
                    """,
                    [selected_user.role, perm, desc],
                )

        return redirect(f"{request.path}?user_id={selected_user.user_id}")

    context = {
        "users": users,
        "selected_user": selected_user,
        "permission_items": permission_items,
        "role_permissions": role_permissions,
    }
    return render(request, "users_access.html", context)
